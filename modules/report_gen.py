"""
Audit Agent: Three-Way Matcher
Module 4 — The Audit Workpaper Report Generator
===============================================

This module renders the structured reconciliation payload from Module 3
(``match_engine.reconcile_ledger_with_invoices``) into a pristine, professionally
styled Excel workbook (.xlsx) that stands on its own as an official
"Substantive Testing Audit Workpaper".

It deliberately drives ``openpyxl`` cell-by-cell rather than ``pandas.to_excel``
so we get full control over fonts, fills, borders, currency number formats,
conditional color-coding, frozen panes, and auto-fitted column widths.

Sheet layout (single tab, "Audit Workpaper"):
    Row 1            : Report title banner.
    Rows 3-4         : "Audit Summary" metadata block (counts).
    Row N (header)   : Deep-navy header row for the main data table.
    Rows N+1 ..      : One styled row per reconciled ledger line, color-coded
                       by audit status.

Public entry point:
    generate_audit_report(reconciliation_data: dict, output_path: str) -> str

Dependencies:
    openpyxl
"""

from __future__ import annotations

import logging
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logger = logging.getLogger("audit_agent.report_gen")
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(
        logging.Formatter("[%(asctime)s] %(name)s %(levelname)s: %(message)s")
    )
    logger.addHandler(_handler)
    logger.setLevel(logging.INFO)


# ---------------------------------------------------------------------------
# Theme / styling constants
# ---------------------------------------------------------------------------
# openpyxl expects ARGB hex strings (no leading '#'); we keep the friendly
# documented hex in comments for traceability against the design spec.
NAVY = "FF1F497D"          # #1F497D  header background
WHITE = "FFFFFFFF"         # white header text
TITLE_BG = "FF17375E"      # slightly darker navy for the title banner
SUMMARY_LABEL_BG = "FFDDEBF7"  # very light blue for the summary labels

# Soft pastel status fills (per spec — no saturated primaries).
FILL_VERIFIED = "FFE2EFDA"        # #E2EFDA soft green       — VERIFIED
FILL_VERIFIED_TOL = "FFD5E8D4"    # #D5E8D4 soft green (deeper) — VERIFIED_WITHIN_TOLERANCE
FILL_EXCEPTION = "FFFFF2CC"       # #FFF2CC soft amber       — EXCEPTION
FILL_MISSING = "FFFCE4D6"         # #FCE4D6 soft red         — MISSING_DOC
FILL_UNRECORDED = "FFE4DFEC"      # #E4DFEC soft purple      — UNRECORDED_INVOICE
FILL_PROCESSING = "FFE7E6E6"      # #E7E6E6 soft gray        — PROCESSING_ERROR / NEEDS_REVIEW
FILL_DUPLICATE = "FFFAC090"       # #FAC090 soft bright orange — POTENTIAL_DUPLICATE_CLAIM

# Accounting-style currency format (per spec).
CURRENCY_FORMAT = r'_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# The main data table columns, in display order. Each entry pairs a header
# label with the result-dict key it reads and a rough "kind" for formatting.
COLUMN_SPEC: list[dict[str, str]] = [
    {"header": "Excel Row", "key": "ledger_row_index", "kind": "int"},
    {"header": "Vendor", "key": "ledger_vendor", "kind": "text"},
    {"header": "Amount", "key": "ledger_amount", "kind": "money"},
    {"header": "Invoice Number", "key": "ledger_invoice_no", "kind": "text"},
    {"header": "Status", "key": "status", "kind": "status"},
    {"header": "Variance", "key": "variance", "kind": "money"},
    {"header": "Matching File", "key": "matching_invoice_file", "kind": "text"},
    {"header": "Audit Notes", "key": "audit_notes", "kind": "notes"},
]

# Map each status to its conditional-formatting fill.
STATUS_FILLS: dict[str, str] = {
    "VERIFIED": FILL_VERIFIED,
    "VERIFIED_WITHIN_TOLERANCE": FILL_VERIFIED_TOL,
    "EXCEPTION": FILL_EXCEPTION,
    "MISSING_DOC": FILL_MISSING,
    "UNRECORDED_INVOICE": FILL_UNRECORDED,
    "PROCESSING_ERROR": FILL_PROCESSING,
    "NEEDS_REVIEW": FILL_PROCESSING,
    "POTENTIAL_DUPLICATE_CLAIM": FILL_DUPLICATE,
}

# The mandatory methodology & limitations disclaimer (review item #8).
DISCLAIMER_TEXT = (
    "METHODOLOGY & LIMITATIONS:  This tool is an automated preparer aid designed "
    "for pre-population. It does not constitute a final assurance conclusion. All "
    "exceptions, errors, and unrecorded liabilities require human auditor validation."
)

# Reusable style objects.
_THIN = Side(style="thin", color="FFBFBFBF")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Column-width bounds so a giant audit note can't blow the sheet out, and a
# short header can't crush the column.
MIN_COL_WIDTH = 10
MAX_COL_WIDTH = 70
NOTES_WRAP_WIDTH = 60  # the "Audit Notes" column wraps rather than runs forever


# ---------------------------------------------------------------------------
# Small styling helpers
# ---------------------------------------------------------------------------
def _set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    number_format: str | None = None,
    border: Border | None = CELL_BORDER,
):
    """
    Write a value to a cell and apply the requested styling in one call.

    Keeps the table-rendering code declarative and free of repetitive
    attribute assignment.

    Returns:
        The styled cell object.
    """
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format
    if border is not None:
        cell.border = border
    return cell


def _auto_fit_columns(ws: Worksheet, max_rows_scanned: int | None = None) -> None:
    """
    Dynamically size each column to its widest cell so nothing clips to '###'.

    Walks every populated cell, tracks the longest rendered string per column
    (respecting newlines for wrapped notes), and sets the column width within
    sensible min/max bounds. The "Audit Notes" column is capped at a wrap width
    since it relies on wrap_text instead of unbounded width.

    Args:
        ws: The worksheet to size.
        max_rows_scanned: Optional cap on how many rows to scan (defaults to all).
    """
    notes_col_idx = next(
        (i for i, c in enumerate(COLUMN_SPEC, start=1) if c["kind"] == "notes"),
        None,
    )

    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        if max_rows_scanned and row[0].row > max_rows_scanned:
            break
        for cell in row:
            if cell.value is None:
                continue
            # For multi-line content, the longest single line drives the width.
            longest_line = max(str(cell.value).split("\n"), key=len)
            widths[cell.column] = max(widths.get(cell.column, 0), len(longest_line))

    for col_idx, raw_width in widths.items():
        # A little padding for breathing room.
        width = raw_width + 3
        if col_idx == notes_col_idx:
            width = min(width, NOTES_WRAP_WIDTH)
        width = max(MIN_COL_WIDTH, min(width, MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------
def _build_title(ws: Worksheet, last_col: int) -> int:
    """
    Render the top title banner across the full table width.

    Returns:
        The next free row index after the banner.
    """
    title_font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_fill = PatternFill("solid", fgColor=TITLE_BG)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = _set_cell(
        ws, 1, 1,
        "Substantive Testing — Three-Way Match Audit Workpaper",
        font=title_font, fill=title_fill, alignment=center, border=None,
    )
    # Paint the merged background across all spanned cells.
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = title_fill
    ws.row_dimensions[1].height = 28
    return 2  # disclaimer banner goes immediately below the title


def _build_disclaimer(ws: Worksheet, last_col: int, start_row: int) -> int:
    """
    Render the mandatory "Methodology & Limitations" disclaimer banner.

    A prominent, wrapped, soft-amber banner spanning the full table width so no
    reader can mistake the automated output for a final assurance conclusion.

    Returns:
        The next free row index after the banner (with a spacer).
    """
    disclaimer_font = Font(name="Calibri", size=10, bold=True, italic=True, color="FF7F6000")
    disclaimer_fill = PatternFill("solid", fgColor="FFFFF2CC")  # soft amber
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFBF9000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=last_col
    )
    _set_cell(
        ws, start_row, 1, "⚠  " + DISCLAIMER_TEXT,
        font=disclaimer_font, fill=disclaimer_fill, alignment=wrap, border=border,
    )
    for col in range(1, last_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = disclaimer_fill
        cell.border = border
    ws.row_dimensions[start_row].height = 42
    return start_row + 2  # spacer row after the banner


def _build_summary(ws: Worksheet, summary: dict, start_row: int) -> int:
    """
    Render the "Audit Summary" metadata block (label/value pairs in a grid).

    Args:
        ws:        Worksheet to write into.
        summary:   The ``summary`` sub-dict from the reconciliation payload.
        start_row: Row at which to begin the block.

    Returns:
        The next free row index after the block (with a spacer).
    """
    heading_font = Font(name="Calibri", size=12, bold=True, color="FF1F497D")
    label_font = Font(name="Calibri", size=10, bold=True, color="FF1F497D")
    value_font = Font(name="Calibri", size=10, bold=True, color="FF000000")
    label_fill = PatternFill("solid", fgColor=SUMMARY_LABEL_BG)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    _set_cell(ws, start_row, 1, "Audit Summary", font=heading_font, border=None)

    # Defensively pull counts (default 0 so a partial payload still renders).
    abs_tol = summary.get("tolerance_absolute", 0.0)
    rel_tol_pct = summary.get("tolerance_relative_pct", 0.0) * 100
    pairs = [
        ("Total Records Tested", summary.get("total_ledger_records", 0)),
        ("Invoices Uploaded", summary.get("total_invoices", 0)),
        ("Verified (exact)", summary.get("verified_count", 0)),
        ("Verified Within Tolerance", summary.get("verified_within_tolerance_count", 0)),
        ("Exceptions Flagged", summary.get("exception_count", 0)),
        ("Missing Documents", summary.get("missing_doc_count", 0)),
        ("Potential Duplicate Claims", summary.get("potential_duplicate_count", 0)),
        ("Unrecorded Invoices (completeness)", summary.get("unrecorded_invoice_count", 0)),
        ("Processing Errors (tooling)", summary.get("processing_error_count", 0)),
        ("Tolerance Applied", f"± ${abs_tol:,.2f}  /  {rel_tol_pct:.3g}%"),
    ]

    row = start_row + 1
    for label, value in pairs:
        _set_cell(ws, row, 1, label, font=label_font, fill=label_fill, alignment=left)
        _set_cell(ws, row, 2, value, font=value_font, alignment=center)
        row += 1

    return row + 1  # spacer row after the block


def _build_table(ws: Worksheet, results: list, start_row: int) -> int:
    """
    Render the main data table: a styled header row plus one row per result.

    Applies currency formatting to money columns, color-codes the Status cell
    by audit outcome, wraps the long notes column, and freezes the panes just
    below the header so the header stays visible while scrolling.

    Args:
        ws:        Worksheet to write into.
        results:   The ``results`` list from the reconciliation payload.
        start_row: Row at which the header should be written.

    Returns:
        The last data row index written.
    """
    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    base_font = Font(name="Calibri", size=10, color="FF000000")
    status_font = Font(name="Calibri", size=10, bold=True, color="FF000000")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    # --- Header row ----------------------------------------------------------
    header_row = start_row
    for col_idx, spec in enumerate(COLUMN_SPEC, start=1):
        _set_cell(
            ws, header_row, col_idx, spec["header"],
            font=header_font, fill=header_fill, alignment=header_align,
        )
    ws.row_dimensions[header_row].height = 24

    # --- Data rows -----------------------------------------------------------
    row = header_row
    for result in results:
        row += 1
        status = str(result.get("status", "")).upper()
        status_fill = (
            PatternFill("solid", fgColor=STATUS_FILLS[status])
            if status in STATUS_FILLS
            else None
        )

        for col_idx, spec in enumerate(COLUMN_SPEC, start=1):
            kind = spec["kind"]
            value = result.get(spec["key"])

            if kind == "money":
                _set_cell(
                    ws, row, col_idx,
                    float(value) if isinstance(value, (int, float)) else 0.0,
                    font=base_font, alignment=right, number_format=CURRENCY_FORMAT,
                )
            elif kind == "int":
                # Invoice-driven rows (UNRECORDED_INVOICE, PROCESSING_ERROR) have
                # no ledger row number — show an em dash instead of a blank cell.
                _set_cell(
                    ws, row, col_idx,
                    "—" if value is None else value,
                    font=base_font, alignment=center,
                )
            elif kind == "status":
                _set_cell(
                    ws, row, col_idx, status,
                    font=status_font, fill=status_fill, alignment=center,
                )
            elif kind == "notes":
                _set_cell(
                    ws, row, col_idx,
                    "" if value is None else str(value),
                    font=base_font, alignment=left_wrap,
                )
            else:  # plain text
                _set_cell(
                    ws, row, col_idx,
                    "" if value is None else str(value),
                    font=base_font, alignment=left,
                )

    # Freeze everything above and including the header so it stays pinned.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    # Enable an auto-filter across the table for auditor convenience.
    last_col_letter = get_column_letter(len(COLUMN_SPEC))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{row}"

    return row


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def generate_audit_report(reconciliation_data: dict, output_path: str) -> str:
    """
    Generate a styled Excel audit workpaper from a reconciliation payload.

    Args:
        reconciliation_data:
            The dict returned by Module 3, containing a ``summary`` block and a
            ``results`` list (see ``match_engine`` for the exact shape).
        output_path:
            Destination path for the .xlsx file. Parent directories are created
            if they do not already exist.

    Returns:
        The absolute path to the workbook that was written.

    Raises:
        TypeError: If ``reconciliation_data`` is not a dict.
        ValueError: If the payload is missing both ``summary`` and ``results``.
    """
    if not isinstance(reconciliation_data, dict):
        raise TypeError("reconciliation_data must be a dict from the match engine.")

    summary = reconciliation_data.get("summary")
    results = reconciliation_data.get("results")
    if summary is None and results is None:
        raise ValueError(
            "reconciliation_data must contain 'summary' and/or 'results' keys."
        )
    summary = summary or {}
    results = results or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Workpaper"
    # Hide gridlines so our borders/fills read as the deliberate design.
    ws.sheet_view.showGridLines = False

    last_col = len(COLUMN_SPEC)

    # Build the sections top-to-bottom: title, disclaimer, summary, data table.
    next_row = _build_title(ws, last_col)
    next_row = _build_disclaimer(ws, last_col, next_row)
    next_row = _build_summary(ws, summary, next_row)
    last_data_row = _build_table(ws, results, next_row)

    # Size columns once everything is on the sheet.
    _auto_fit_columns(ws)

    # Ensure the destination directory exists, then save.
    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)

    try:
        wb.save(output_path)
    except PermissionError as exc:
        # Most common real-world failure: the file is open in Excel.
        raise PermissionError(
            f"Could not write '{output_path}'. Is the file open in Excel? ({exc})"
        ) from exc

    abs_path = os.path.abspath(output_path)
    logger.info(
        "Audit workpaper written to %s (%d data rows).", abs_path, last_data_row
    )
    return abs_path


# ---------------------------------------------------------------------------
# Isolated self-test: build a sample workpaper from a mock Module 3 payload.
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    logging.getLogger("audit_agent.report_gen").setLevel(logging.INFO)

    # Mock payload mirroring exactly what the upgraded match_engine.py emits —
    # one row per enterprise status so the styling is exercised end-to-end.
    mock_reconciliation = {
        "summary": {
            "total_ledger_records": 5,
            "total_invoices": 7,
            "verified_count": 1,
            "verified_within_tolerance_count": 1,
            "exception_count": 1,
            "missing_doc_count": 1,
            "potential_duplicate_count": 1,
            "unrecorded_invoice_count": 1,
            "processing_error_count": 1,
            "tolerance_absolute": 0.05,
            "tolerance_relative_pct": 0.005,
        },
        "results": [
            {"ledger_row_index": 2, "ledger_vendor": "Acme Corp",
             "ledger_amount": 1250.00, "ledger_invoice_no": "INV-001",
             "status": "VERIFIED", "variance": 0.0,
             "matching_invoice_file": "acme_inv_001.pdf",
             "audit_notes": "VERIFIED: amount ties to the penny; invoice no. and vendor match."},
            {"ledger_row_index": 3, "ledger_vendor": "Globex LLC",
             "ledger_amount": 500.03, "ledger_invoice_no": "INV-002",
             "status": "VERIFIED_WITHIN_TOLERANCE", "variance": 0.03,
             "matching_invoice_file": "globex_inv_002.pdf",
             "audit_notes": "VERIFIED WITHIN TOLERANCE: $0.03 variance within ±$2.50 (likely tax/rounding)."},
            {"ledger_row_index": 4, "ledger_vendor": "Wayne Ent",
             "ledger_amount": 750.00, "ledger_invoice_no": "INV-003",
             "status": "EXCEPTION", "variance": -150.0,
             "matching_invoice_file": "wayne_inv_003.pdf",
             "audit_notes": "EXCEPTION: amount variance of -150.00 exceeds the allowed tolerance."},
            {"ledger_row_index": 5, "ledger_vendor": "Initech",
             "ledger_amount": 3400.00, "ledger_invoice_no": "INV-404",
             "status": "MISSING_DOC", "variance": 3400.0,
             "matching_invoice_file": None,
             "audit_notes": "No supporting invoice document could be matched. Full amount unverified."},
            {"ledger_row_index": 6, "ledger_vendor": "Acme Corp",
             "ledger_amount": 1250.00, "ledger_invoice_no": "INV-001",
             "status": "POTENTIAL_DUPLICATE_CLAIM", "variance": 0.0,
             "matching_invoice_file": "acme_inv_001.pdf",
             "audit_notes": "POTENTIAL DUPLICATE CLAIM: invoice INV-001 already claimed by ledger row 2."},
            {"ledger_row_index": None, "ledger_vendor": "Umbrella Inc",
             "ledger_amount": 9999.99, "ledger_invoice_no": "INV-777",
             "status": "UNRECORDED_INVOICE", "variance": 9999.99,
             "matching_invoice_file": "umbrella_777.pdf",
             "audit_notes": "COMPLETENESS RISK: invoice not matched to any ledger line — potential unrecorded liability."},
            {"ledger_row_index": None, "ledger_vendor": "(unreadable)",
             "ledger_amount": 0.0, "ledger_invoice_no": "",
             "status": "PROCESSING_ERROR", "variance": 0.0,
             "matching_invoice_file": "blurry_scan.pdf",
             "audit_notes": "AI EXTRACTION FAILURE (HTTP 429). Tooling limitation, NOT a missing document."},
        ],
    }

    sample_path = os.path.join(os.path.dirname(__file__), "test_audit_workpaper.xlsx")
    written = generate_audit_report(mock_reconciliation, sample_path)

    print(f"\n[self-test] Sample workpaper generated at:\n  {written}")

    # --- Lightweight verification by re-opening the file ---------------------
    from openpyxl import load_workbook

    wb_check = load_workbook(written)
    ws_check = wb_check["Audit Workpaper"]

    # The methodology disclaimer must be present near the top of the sheet.
    top_text = " ".join(
        str(ws_check.cell(row=r, column=1).value or "")
        for r in range(1, 6)
    )
    assert "does not constitute a final assurance conclusion" in top_text, (
        "Methodology & Limitations disclaimer banner must be present."
    )

    # Find the header row by locating the "Status" header label.
    status_col = header_row = None
    for r in range(1, 30):
        for c in range(1, len(COLUMN_SPEC) + 1):
            if ws_check.cell(row=r, column=c).value == "Status":
                header_row, status_col = r, c
                break
        if header_row:
            break
    assert header_row is not None, "Header row with 'Status' must exist."

    # Every status must carry its mapped pastel fill.
    seen_statuses = []
    r = header_row + 1
    while ws_check.cell(row=r, column=status_col).value:
        cell = ws_check.cell(row=r, column=status_col)
        status_val = cell.value
        seen_statuses.append(status_val)
        actual = cell.fill.fgColor.rgb
        assert actual == STATUS_FILLS[status_val], (
            f"Row {r} status '{status_val}' fill {actual} != {STATUS_FILLS[status_val]}"
        )
        r += 1

    assert seen_statuses == [
        "VERIFIED", "VERIFIED_WITHIN_TOLERANCE", "EXCEPTION", "MISSING_DOC",
        "POTENTIAL_DUPLICATE_CLAIM", "UNRECORDED_INVOICE", "PROCESSING_ERROR",
    ], f"Unexpected status ordering: {seen_statuses}"

    # Spot-check the currency format on the first Amount cell.
    amount_col = next(
        i for i, s in enumerate(COLUMN_SPEC, start=1) if s["key"] == "ledger_amount"
    )
    amt_cell = ws_check.cell(row=header_row + 1, column=amount_col)
    assert amt_cell.number_format == CURRENCY_FORMAT, "Amount must use currency format."
    assert abs(amt_cell.value - 1250.00) < 1e-9, "Amount value must round-trip."

    # Invoice-driven rows must show an em-dash for the (absent) Excel row number.
    row_col = next(
        i for i, s in enumerate(COLUMN_SPEC, start=1) if s["key"] == "ledger_row_index"
    )
    unrecorded_excel_row = header_row + 6  # 6th data row = UNRECORDED_INVOICE
    assert ws_check.cell(row=unrecorded_excel_row, column=row_col).value == "—", (
        "Invoice-driven rows must render '—' for the missing ledger row index."
    )

    print("[self-test] Disclaimer, all 7 status colors, currency & layout verified. [OK]")
    print("[self-test] Open the file in Excel to view the styled workpaper.")
